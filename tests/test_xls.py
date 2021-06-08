import unittest
from pathlib import Path

from inveniautils.xlsutil import Workbook
import openpyxl
import xlrd  # type: ignore

xls_file_name = "Sample_XLS.xls"
xlsx_file_name = "Sample_XLSX.xlsx"
sample_xls_path = Path("tests") / "files" / xls_file_name
sample_xlsx_path = Path("tests") / "files" / xlsx_file_name


class TestXLSUtil(unittest.TestCase):
    def setup_xls(self):
        file = open(sample_xls_path, "rb")
        content = file.read()
        self.assertNotEqual(content, None)
        return Workbook(content, xls_file_name)

    def setup_xlsx(self):
        file = open(sample_xlsx_path, "rb")
        content = file.read()
        self.assertNotEqual(content, None)
        return Workbook(content, xlsx_file_name)

    def test_no_filename_xls(self):
        file = open(sample_xls_path, "rb")
        content = file.read()
        self.assertNotEqual(content, None)
        workbook = Workbook(content)
        self.assertEqual(workbook.is_xls, True)

    def test_no_filename_xlsx(self):
        file = open(sample_xlsx_path, "rb")
        content = file.read()
        self.assertNotEqual(content, None)
        workbook = Workbook(content)
        self.assertEqual(workbook.is_xls, False)

    def test_xls_sheet_names(self):
        workbook = self.setup_xls()
        sheet_names = workbook.sheet_names()
        self.assertEqual(len(sheet_names), 1)
        self.assertEqual(sheet_names[0], "TestSheet")

    def test_xlsx_sheet_names(self):
        workbook = self.setup_xlsx()
        sheet_names = workbook.sheet_names()
        self.assertEqual(len(sheet_names), 1)
        self.assertEqual(sheet_names[0], "TestSheet")

    def test_xls_sheet_by_name(self):
        workbook = self.setup_xls()
        sheet = workbook.sheet_by_name("TestSheet")
        self.assertEqual(sheet.name, "TestSheet")

    def test_xlsx_sheet_by_name(self):
        workbook = self.setup_xlsx()
        sheet = workbook.sheet_by_name("TestSheet")
        self.assertEqual(sheet.name, "TestSheet")

    def test_xls_row(self):
        expected_0 = ["number", "name", "value"]
        expected_1 = [1, "A", 3]
        workbook = self.setup_xls()
        sheet = workbook.sheet_by_name("TestSheet")
        row_0 = sheet.row(0)
        row_1 = sheet.row(1)

        for i in range(0, 2):
            self.assertEqual(row_0[i].value, expected_0[i])
            self.assertEqual(row_1[i].value, expected_1[i])

    def test_xlsx_row(self):
        expected_0 = ["number", "name", "value"]
        expected_1 = [1, "A", 3]
        workbook = self.setup_xlsx()
        sheet = workbook.sheet_by_name("TestSheet")
        row_0 = sheet.row(0)
        row_1 = sheet.row(1)

        for i in range(0, 2):
            self.assertEqual(row_0[i].value, expected_0[i])
            self.assertEqual(row_1[i].value, expected_1[i])

    def test_xls_get_cell_type(self):
        workbook = self.setup_xls()
        sheet = workbook.sheet_by_name("TestSheet")
        row = sheet.row(1)

        # get the type mappings
        number_cell = row[0].NUMBER_CELL
        text_cell = row[0].TEXT_CELL
        empty_cell = row[0].EMPTY_CELL

        # check the type mappings
        self.assertEqual(number_cell, xlrd.XL_CELL_NUMBER)
        self.assertEqual(text_cell, xlrd.XL_CELL_TEXT)
        self.assertEqual(empty_cell, xlrd.XL_CELL_EMPTY)

        # make sure each cell is the correct type
        self.assertEqual(row[0]._get_cell_type(), number_cell)
        self.assertEqual(row[1]._get_cell_type(), text_cell)
        self.assertEqual(row[2]._get_cell_type(), number_cell)

    def test_xlsx_get_cell_type(self):
        workbook = self.setup_xlsx()
        sheet = workbook.sheet_by_name("TestSheet")
        row = sheet.row(1)

        # get the type mappings
        number_cell = row[0].NUMBER_CELL
        text_cell = row[0].TEXT_CELL
        empty_cell = row[0].EMPTY_CELL

        # check the type mappings
        self.assertEqual(number_cell, openpyxl.cell.cell.TYPE_NUMERIC)
        self.assertEqual(text_cell, openpyxl.cell.cell.TYPE_STRING)
        self.assertEqual(empty_cell, openpyxl.cell.cell.TYPE_NULL)

        # make sure each cell is the correct type
        self.assertEqual(row[0]._get_cell_type(), number_cell)
        self.assertEqual(row[1]._get_cell_type(), text_cell)
        self.assertEqual(row[2]._get_cell_type(), number_cell)

    def test_xls_cell(self):
        expected_0 = ["number", "name", "value"]
        expected_1 = [1, "A", 3]
        workbook = self.setup_xls()
        sheet = workbook.sheet_by_name("TestSheet")

        for i in range(0, 2):
            self.assertEqual(sheet.cell(0, i).value, expected_0[i])
            self.assertEqual(sheet.cell(1, i).value, expected_1[i])

    def test_xlsx_cell(self):
        expected_0 = ["number", "name", "value"]
        expected_1 = [1, "A", 3]
        workbook = self.setup_xlsx()
        sheet = workbook.sheet_by_name("TestSheet")

        for i in range(0, 2):
            self.assertEqual(sheet.cell(0, i).value, expected_0[i])
            self.assertEqual(sheet.cell(1, i).value, expected_1[i])
