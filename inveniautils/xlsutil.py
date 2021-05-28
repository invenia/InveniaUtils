from io import BytesIO

import openpyxl
import xlrd  # type: ignore


class Cell:
    # OPENPYXL data_type mappings are found here:
    # https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/cell/cell.html
    EMPTY_CELL_OPENPYXL = openpyxl.cell.cell.TYPE_NULL
    TEXT_CELL_OPENPYXL = openpyxl.cell.cell.TYPE_STRING
    NUMBER_CELL_OPENPYXL = openpyxl.cell.cell.TYPE_NUMERIC

    # XLRD ctype mappings are found here: https://pythonhosted.org/xlrd3/cell.html
    EMPTY_CELL_XLRD = xlrd.XL_CELL_EMPTY
    TEXT_CELL_XLRD = xlrd.XL_CELL_TEXT
    NUMBER_CELL_XLRD = xlrd.XL_CELL_NUMBER

    def __init__(self, is_xls: bool, cell):
        self.is_xls = is_xls
        self.cell = cell
        self.ctype = self._get_cell_type()
        self.value = self.cell.value
        self._set_type_mappings()

    def _get_cell_type(self):
        return self.cell.ctype if self.is_xls else self.cell.data_type

    def _set_type_mappings(self):
        if self.is_xls:
            self.EMPTY_CELL = self.EMPTY_CELL_XLRD
            self.TEXT_CELL = self.TEXT_CELL_XLRD
            self.NUMBER_CELL = self.NUMBER_CELL_XLRD
        else:
            self.EMPTY_CELL = self.EMPTY_CELL_OPENPYXL
            self.TEXT_CELL = self.TEXT_CELL_OPENPYXL
            self.NUMBER_CELL = self.NUMBER_CELL_OPENPYXL


class Sheet:
    def __init__(self, is_xls: bool, sheet):
        self.is_xls = is_xls
        self.sheet = sheet
        self.nrows = self._set_num_rows()
        self.ncols = self._set_num_cols()
        self.name = self._set_name()

    def _set_num_rows(self):
        return self.sheet.nrows if self.is_xls else self.sheet.max_row

    def _set_num_cols(self):
        return self.sheet.ncols if self.is_xls else self.sheet.max_column

    def _set_name(self):
        return self.sheet.name if self.is_xls else self.sheet.title

    def cell(self, row, column):
        # openpyxl starts indexing at 1
        return (
            self.sheet.cell(row, column)
            if self.is_xls
            else self.sheet.cell(row + 1, column + 1)
        )

    def row(self, row_num: int):
        # openpyxl starts indexing at 1
        initial_row = (
            self.sheet.row(row_num) if self.is_xls else self.sheet[row_num + 1]
        )

        return list(Cell(self.is_xls, item) for item in initial_row)

    def row_values(self, row_num: int):
        # openpyxl starts indexing at 1
        return (
            self.sheet.row_values(row_num)
            if self.is_xls
            else list(item.value for item in self.sheet[row_num + 1])
        )


class Workbook:
    def __init__(self, filename: str, content):
        self.is_xls = filename.endswith(".xls")
        self.workbook = self._open_workbook(content)

    def _open_workbook(self, content):
        return (
            xlrd.open_workbook(file_contents=content)
            if self.is_xls
            else openpyxl.load_workbook(BytesIO(content))
        )

    def sheet_names(self):
        return self.workbook.sheet_names() if self.is_xls else self.workbook.sheetnames

    def sheet_by_name(self, sheet_name: str):
        sheet = (
            self.workbook.sheet_by_name(sheet_name)
            if self.is_xls
            else self.workbook[sheet_name]
        )
        return Sheet(self.is_xls, sheet)
